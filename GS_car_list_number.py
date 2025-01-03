import os
import pandas as pd
import glob
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from collections import Counter, defaultdict
from GS_vehicle_dict import vehicle_dict

# ---------------------------------------------------------------------------------
# 1) 기본 설정
# ---------------------------------------------------------------------------------
base_path = r"D:/SamsungSTF/Data/GSmbiz/BMS_Data"    # BMS 폴더 최상위 경로
gps_base_path = r"D:/SamsungSTF/Data/GSmbiz/gps_altitude"  # 새로 추가한 gps_altitude 폴더 최상위 경로

def format_device_number(device_number):
    """ 단말기번호를 000-0000-0000 형태로 변환 """
    return f"{device_number[:3]}-{device_number[3:7]}-{device_number[7:]}"

# ---------------------------------------------------------------------------------
# 2) 모든 단말기번호 수집
# ---------------------------------------------------------------------------------
all_device_numbers = set()
for device_numbers in vehicle_dict.values():
    all_device_numbers.update(device_numbers)

# ---------------------------------------------------------------------------------
# 3) 결과 저장할 구조 준비
# ---------------------------------------------------------------------------------
results = []
monthly_counts = defaultdict(lambda: defaultdict(int))  # 차종별 / 연월별 파일 존재 카운트

# ---------------------------------------------------------------------------------
# 4) 디렉토리 구조 탐색 및 파일 존재 여부 확인
# ---------------------------------------------------------------------------------
for device_number in all_device_numbers:
    # 4-1) 해당 단말기의 차종(car_model) 확인
    car_model = "Unknown"
    for model, numbers in vehicle_dict.items():
        if device_number in numbers:
            car_model = model
            break

    # 4-2) record에 기본 정보 입력
    record = {
        "단말기번호": format_device_number(device_number),
        "차종": car_model
    }
    collected_periods = 0  # 수집된 기간(월)의 개수

    # 4-3) 2023 ~ 2024년(2025년은 0월이면 break) 루프
    for year in range(2023, 2025):  # 2023, 2024
        for month in range(1, 13):  # 1 ~ 12월
            if year == 2025 and month > 0:
                break  # 2024년 12월까지만 확인한다고 가정

            # 연월 문자열
            year_month = f"{year}-{month:02d}"

            # bms, bms_altitude 폴더 내 패턴 설정
            folder_path = os.path.join(base_path, device_number, year_month)
            bms_pattern = os.path.join(folder_path, f"bms_{device_number}_{year_month}-*.csv")
            bms_altitude_pattern = os.path.join(folder_path, f"bms_altitude_{device_number}_{year_month}-*.csv")

            # 추가: gps_altitude 폴더 내 파일 확인
            gps_folder_path = os.path.join(gps_base_path, device_number, year_month)
            gps_altitude_pattern = os.path.join(gps_folder_path, f"gps_altitude_{device_number}_{year}-{month:02d}-*.csv")

            # 실제 파일 존재 여부 확인
            bms_files = glob.glob(bms_pattern)
            bms_altitude_files = glob.glob(bms_altitude_pattern)
            gps_altitude_files = glob.glob(gps_altitude_pattern)

            # 파일 존재 유무에 따른 status (단순화 예시)
            # ※ 필요하다면 여러 조건(동시에 존재, 일부만 존재 등)을 더 세밀히 구분 가능
            if bms_files and not bms_altitude_files and not gps_altitude_files:
                status = "bms_파일만 존재"  # 고도정보 없는 bms
            elif bms_altitude_files and not gps_altitude_files:
                status = "bms_altitude_파일 존재"  # altitude 정보 포함된 bms
            elif gps_altitude_files:
                status = "gps_altitude_파일 존재"  # GPS와 BMS가 동시에 존재한다고 가정
            else:
                status = "파일 없음"

            file_count = 0
            if os.path.exists(folder_path):
                file_count = len(os.listdir(folder_path))
            # gps 폴더 쪽 파일도 카운트에 합산하고 싶다면 아래 식처럼 추가 가능
            # if os.path.exists(gps_folder_path):
            #     file_count += len(os.listdir(gps_folder_path))

            # 엑셀에 들어갈 컬럼명 (YY-MM 형태)
            column_name = f"{str(year)[-2:]}-{month:02d}"
            # 상태와 파일 개수를 표시 (또는 필요시 file_count+gps_count 등 합산도 가능)
            record[column_name] = (
                f"{status} ({file_count})"
                if (file_count > 0 or gps_altitude_files)  # gps만 있더라도 존재로 판단
                else None
            )

            # 해당 연월에 파일(어떤 것이든) 존재한다고 간주하면 collected_periods 증가
            if file_count > 0 or gps_altitude_files:
                collected_periods += 1
                monthly_counts[car_model][column_name] += 1

    # "기간" 열에 수집된 기간(월) 개수 기입
    record["기간"] = collected_periods
    results.append(record)

# ---------------------------------------------------------------------------------
# 5) DataFrame 생성 및 정렬
# ---------------------------------------------------------------------------------
df = pd.DataFrame(results)
df.sort_values(by="단말기번호", ascending=True, inplace=True)

# ---------------------------------------------------------------------------------
# 6) 차종별 차량 개수 계산
# ---------------------------------------------------------------------------------
car_model_counts = Counter(df["차종"])

# ---------------------------------------------------------------------------------
# 7) 엑셀로 저장 (기본 데이터 시트)
# ---------------------------------------------------------------------------------
output_path = os.path.join(os.path.dirname(base_path), "단말기_번호별_차종.xlsx")
df.to_excel(output_path, index=False)

# ---------------------------------------------------------------------------------
# 8) 엑셀 스타일링 (색상, 테두리, 열 너비, etc.)
# ---------------------------------------------------------------------------------
wb = openpyxl.load_workbook(output_path)
ws = wb.active

# 8-1) "차종별 차량 개수" 시트 추가
summary_ws = wb.create_sheet(title="차종별_차량_개수")
header_row = ["차종", "차량 개수"]
year_month_list = []
for year in range(2023, 2025):
    for month in range(1, 13):
        if year == 2025 and month > 0:
            break
        year_month_list.append(f"{str(year)[-2:]}-{month:02d}")
header_row += year_month_list
summary_ws.append(header_row)

for model, count in car_model_counts.items():
    row = [model, count]
    for ym in year_month_list:
        row.append(monthly_counts[model].get(ym, 0))
    summary_ws.append(row)

# 8-2) Fill(배경색) 정의
green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 연한 초록
blue_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")   # 연한 보라(파랑에 가까운 보라)
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") # 주황색
yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # 연한 노랑
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")    # 연한 빨강
light_salmon_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid") # 살몬색
light_pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")   # 연한 분홍
pink_fill = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")   # 연한 핑크

# 8-3) Border(테두리) & Alignment(정렬) 정의
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
center_alignment = Alignment(horizontal='center', vertical='center')

# 8-4) 데이터 셀 스타일 설정
#     (BMS, bms_altitude, gps_altitude 파일 존재 여부에 따라 셀 배경색 표시)
for row in ws.iter_rows(min_row=2,  # 제목행 제외
                        min_col=3,  # '단말기번호'(1열), '차종'(2열) 제외
                        max_col=ws.max_column-1,  # 마지막 열 '기간' 제외
                        max_row=ws.max_row):
    for cell in row:
        if cell.value:
            cell_str = str(cell.value)
            if "bms_파일만 존재" in cell_str:
                # 고도정보 없는 BMS -> 초록색
                cell.fill = green_fill
                # ( ... ) 안에 파일 개수 정수만 추출
                file_count = int(cell_str.split()[-1].strip('()'))
                cell.value = file_count

            elif "bms_altitude_파일 존재" in cell_str:
                # altitude 정보만 존재하는 BMS -> 보라색
                cell.fill = blue_fill
                file_count = int(cell_str.split()[-1].strip('()'))
                cell.value = file_count

            elif "gps_altitude_파일 존재" in cell_str:
                # GPS와 BMS 둘 다 존재 -> 주황색
                cell.fill = orange_fill
                # bms 폴더 쪽 파일 개수만 추출(필요에 따라 수정 가능)
                file_count = 0
                # 괄호 안의 숫자 있는 경우
                if "(" in cell_str:
                    file_count = int(cell_str.split()[-1].strip('()'))
                cell.value = file_count

        # 중앙 정렬
        cell.alignment = center_alignment

# 8-5) 단말기번호(A열), 차종(B열) 노란색 배경
for row in ws.iter_rows(min_row=2,
                        min_col=1,
                        max_col=2,
                        max_row=ws.max_row):
    for cell in row:
        cell.fill = yellow_fill
        cell.alignment = center_alignment

# 8-6) 마지막 열(기간)에 대한 스타일 (예시: 기간=12 -> 살몬색 등)
for row in ws.iter_rows(min_row=2,
                        min_col=ws.max_column,
                        max_col=ws.max_column,
                        max_row=ws.max_row):
    for cell in row:
        if cell.value == 12:
            cell.fill = light_salmon_fill
        elif 7 < cell.value < 12:
            cell.fill = red_fill
        elif 3 < cell.value <= 7:
            cell.fill = light_pink_fill
        else:
            cell.fill = pink_fill
        cell.alignment = center_alignment

# 8-7) 모든 셀에 테두리 적용
for row in ws.iter_rows(min_row=1,
                        min_col=1,
                        max_col=ws.max_column,
                        max_row=ws.max_row):
    for cell in row:
        cell.border = border

# 8-8) 열 너비 자동 조정(단순 추정)
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            length = len(str(cell.value))
            if length > max_length:
                max_length = length
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

# 차종 열(컬럼 B) 너비를 조금 더 넓게 설정
ws.column_dimensions['B'].width = ws.column_dimensions['B'].width * 1.5

# ---------------------------------------------------------------------------------
# 9) [색상 범례] 병합 셀 추가
# ---------------------------------------------------------------------------------
legend_row = ws.max_row + 2  # 데이터가 끝난 뒤 2줄 아래에 범례 삽입

# 시트 전체 열을 하나의 셀로 병합
ws.merge_cells(
    start_row=legend_row,
    start_column=1,
    end_row=legend_row+3,
    end_column=ws.max_column
)
legend_cell = ws.cell(row=legend_row, column=1)
legend_cell.value = (
    "[색상 범례]\n"
    " - 초록색: 고도정보 없는 BMS 파일 (bms_파일만 존재)\n"
    " - 보라색: altitude 정보만 존재하는 BMS 파일 (bms_altitude_파일 존재)\n"
    " - 주황색: GPS와 BMS 파일이 모두 존재 (gps_altitude_파일 존재)"
)
legend_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 범례 셀 테두리 (병합범위 전체)
for row in ws.iter_rows(
    min_row=legend_row, max_row=legend_row+3,
    min_col=1, max_col=ws.max_column
):
    for cell in row:
        cell.border = border

# ---------------------------------------------------------------------------------
# 10) 저장
# ---------------------------------------------------------------------------------
wb.save(output_path)
print("차종별 차량 개수 및 파일 존재 여부 확인 완료, 범례 추가 후 저장 완료.")
